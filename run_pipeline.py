"""
Backlink Outreach Pipeline
===========================
Runs weekly (or on demand) via GitHub Actions.

Flow:
  1. Pull competitor referring domains from Ahrefs API
  2. Pull target domain's existing referring domains
  3. Load existing publisher database (database.csv) as blocklist
  4. Gap analysis — in competitors but NOT in target and NOT in database
  5. Filter: DR 10+, Traffic 1000+, not spam
  6. Score topical relevance against accepted topic list
  7. Scrape emails + phones from contact/advertise pages
  8. Output Excel file to /output/

Environment variables required:
  AHREFS_API_KEY  — your Ahrefs API v3 key
"""

import os, re, csv, time, random, json, requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════

AHREFS_API_KEY = os.environ.get("AHREFS_API_KEY", "")
TARGET_DOMAIN  = "32red.com"
OUTPUT_DIR     = "output"

# Path to your existing publisher database CSV (exported from Google Sheets).
# Domains in this file are treated as already-known and excluded from outreach.
# Upload a fresh export whenever your team updates the sheet.
DATABASE_CSV   = "database.csv"

# Top 10 competitors (from Ahrefs Competing Domains, sorted by common keywords)
COMPETITORS = [
    "paddypower.com",
    "betfair.com",
    "unibet.co.uk",
    "williamhill.com",
    "betvictor.com",
    "coral.co.uk",
    "mrq.com",
    "grosvenorcasinos.com",
    "netbet.co.uk",
    "ladbrokes.com",
]

# Rows to pull per competitor (increase for larger lists — costs more Ahrefs units)
ROWS_PER_COMPETITOR = 500   # 500 × 10 = 5,000 domains before deduplication
ROWS_TARGET_DOMAIN  = 2000  # Pull existing 32red links to exclude

# Quality filters
MIN_DR      = 10
MIN_TRAFFIC = 1000

# Scraping
SCRAPE_WORKERS   = 8        # Concurrent scraper threads
REQUEST_TIMEOUT  = 8        # Seconds per HTTP request
DELAY_MIN        = 0.3      # Polite delay between requests (seconds)
DELAY_MAX        = 0.8

CONTACT_PATHS = [
    "/contact", "/contact-us", "/contact_us",
    "/advertise", "/advertise-with-us", "/advertising",
    "/write-for-us", "/write-for-us/", "/contribute",
    "/partnerships", "/sponsored", "/about", "/about-us",
    "/press", "/media", "/work-with-us",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
}

# Accepted topic categories + their keyword signals
TOPIC_KEYWORDS = {
    "Gambling":       ["casino","gambling","betting","slots","poker","bingo","jackpot","wagering","bookmaker","sportsbook","punter","odds","free spins","igaming"],
    "Sports":         ["sport","football","soccer","cricket","rugby","tennis","golf","boxing","athletics","formula 1","motorsport","nfl","nba","mlb","nhl","cycling","swimming"],
    "Tech":           ["technology","tech","software","gadget","digital","app","coding","developer","artificial intelligence","cybersecurity","startup","saas","hardware"],
    "Gaming/eSports": ["gaming","esports","e-sports","gamer","twitch","streamer","video game","pc gaming","console","playstation","xbox","nintendo","overwatch","league of legends"],
    "Travel":         ["travel","holiday","vacation","tourism","destination","hotel","flight","backpack","adventure","cruise","resort","passport"],
    "Finance":        ["finance","money","investment","trading","stock","forex","financial","wealth","budget","pension","mortgage","insurance","banking","economy"],
    "Crypto":         ["crypto","bitcoin","blockchain","ethereum","defi","nft","web3","altcoin","binance","coinbase","decentralized","token"],
    "General News":   ["news","breaking","journalist","reporter","media","press","editorial","daily","weekly","bulletin","current affairs"],
    "Global News":    ["world news","international","global","foreign affairs","geopolitics","reuters","associated press"],
    "Entertainment":  ["entertainment","celebrity","lifestyle","culture","music","film","tv","television","movie","series","award","fashion"],
    "Magazines":      ["magazine","editorial","feature","publication","issue","subscriber","print","glossy","monthly"],
}

NOISE_EMAIL_DOMAINS = {
    "example.com","domain.com","youremail.com","sentry.io","wixpress.com",
    "schema.org","w3.org","google.com","apple.com","facebook.com","twitter.com",
    "cloudflare.com","wordpress.com","jquery.com","amazonaws.com","datatables.net",
}

# Generic/infrastructure domains to always exclude from results
DOMAIN_BLOCKLIST = {
    "google.com","youtube.com","twitter.com","x.com","facebook.com","instagram.com",
    "linkedin.com","pinterest.com","reddit.com","wikipedia.org","apple.com",
    "microsoft.com","amazonaws.com","cloudfront.net","github.io","github.com",
    "wordpress.com","blogspot.com","tumblr.com","medium.com","substack.com",
    "bit.ly","tinyurl.com","bitly.com","goo.gl","t.co","t.me","ift.tt",
    "squarespace.com","wix.com","wixsite.com","weebly.com","jimdo.com",
    "jimdofree.com","strikingly.com","netlify.app","netlify.com","vercel.app",
    "herokuapp.com","azurewebsites.net","azurefd.net","secureserver.net",
    "hostingersite.com","bigcartel.com","homestead.com",
    "hatena.ne.jp","ameblo.jp","hatenablog.com","livedoor.jp","fc2.com",
    "rakuten.co.jp","yahoo.co.jp","biglobe.ne.jp","nifty.com","exblog.jp",
    "sakura.ne.jp","xsrv.jp","goo.ne.jp","jugem.jp","main.jp","jp.net",
    "livejournal.com","zendesk.com","trello.com","atlassian.net","crunchbase.com",
    "slideshare.net","prweb.com","prnewswire.com","globenewswire.com",
    "us.com","uk.com","de.com","it.com","in.net","com.de",
    "ovh.net","free.fr","haendlerbund.de","csdn.net","163.com","sina.com.cn",
    "quantcast.com","snowplow.io","onetrust.com","vmware.com","dnb.com",
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.I)
PHONE_RE = re.compile(
    r"(?:\+44|0044|0)[\s\-\.]?(?:\d[\s\-\.]?){9,10}"
    r"|(?:\+1[\s\-\.]?)?\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]?\d{4}"
)


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE BLOCKLIST
# ══════════════════════════════════════════════════════════════════════════════

def load_database_domains() -> set:
    """
    Read database.csv (your existing publisher database exported from Google Sheets)
    and return a set of domains to exclude from outreach.
    If the file doesn't exist, returns an empty set and continues normally.
    """
    if not os.path.exists(DATABASE_CSV):
        print(f"  [info] No database.csv found — skipping database filter.")
        return set()

    domains = set()
    with open(DATABASE_CSV, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        # Normalise headers (strip whitespace + newlines from multi-line headers)
        reader.fieldnames = [h.replace("\n", " ").strip() for h in reader.fieldnames]
        for row in reader:
            row = {k.replace("\n", " ").strip(): v for k, v in row.items()}
            domain = row.get("Domain", "").strip().lower()
            if domain:
                domains.add(domain)

    print(f"  [database] Loaded {len(domains)} existing domains to exclude")
    return domains


# ══════════════════════════════════════════════════════════════════════════════
# AHREFS API
# ══════════════════════════════════════════════════════════════════════════════

def ahrefs_get(endpoint, params):
    """Call Ahrefs API v3 and return parsed JSON."""
    url = f"https://api.ahrefs.com/v3/{endpoint}"
    headers = {
        "Authorization": f"Bearer {AHREFS_API_KEY}",
        "Accept": "application/json",
    }
    resp = requests.get(url, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def get_referring_domains(domain, limit=500):
    """Pull referring domains for a given domain from Ahrefs."""
    print(f"  → Ahrefs: pulling {limit} referring domains for {domain}")
    params = {
        "target": domain,
        "mode": "subdomains",
        "select": "domain,domain_rating,traffic_domain,is_spam",
        "where": json.dumps({
            "and": [
                {"field": "domain_rating",  "is": ["gte", MIN_DR]},
                {"field": "traffic_domain", "is": ["gte", MIN_TRAFFIC]},
                {"field": "is_spam",        "is": ["eq", 0]},
            ]
        }),
        "order_by": "domain_rating:desc",
        "limit": limit,
    }
    data = ahrefs_get("site-explorer/referring-domains", params)
    return data.get("refdomains", [])


# ══════════════════════════════════════════════════════════════════════════════
# GAP ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def build_gap_list():
    """
    Pull competitor + target referring domains, return gap domains
    (in competitors but NOT already linking to the target, and NOT in the database).
    """
    print("\n[1/5] Loading existing publisher database...")
    database_domains = load_database_domains()

    print(f"\n[2/5] Pulling {TARGET_DOMAIN} existing referring domains...")
    target_domains = {
        r["domain"]
        for r in get_referring_domains(TARGET_DOMAIN, limit=ROWS_TARGET_DOMAIN)
    }
    print(f"      {TARGET_DOMAIN} has {len(target_domains)} qualifying referring domains")

    # Combined exclusion set: existing links + database + generic blocklist
    exclude = target_domains | database_domains | DOMAIN_BLOCKLIST
    print(f"      Total exclusions: {len(exclude)} domains")

    print("\n[3/5] Pulling competitor referring domains...")
    competitor_map = {}   # domain -> {domain_rating, traffic_domain}
    for comp in COMPETITORS:
        rows = get_referring_domains(comp, limit=ROWS_PER_COMPETITOR)
        for r in rows:
            d = r["domain"]
            if d not in competitor_map:
                competitor_map[d] = {
                    "domain":          d,
                    "domain_rating":   r["domain_rating"],
                    "traffic_domain":  r["traffic_domain"],
                    "competitors_found_on": [],
                }
            competitor_map[d]["competitors_found_on"].append(comp)
        print(f"      {comp}: {len(rows)} domains pulled")
        time.sleep(0.5)

    # Gap = in competitors, NOT in exclusion set (target + database + blocklist)
    gap = [
        v for k, v in competitor_map.items()
        if k not in exclude
    ]

    # Sort by DR desc
    gap.sort(key=lambda x: x["domain_rating"], reverse=True)

    print(f"\n      Gap list: {len(gap)} unique domains after filtering")
    return gap


# ══════════════════════════════════════════════════════════════════════════════
# TOPIC RELEVANCE SCORING
# ══════════════════════════════════════════════════════════════════════════════

def score_topic(domain, title="", description=""):
    """
    Score a domain against accepted topic keywords.
    Returns (best_topic, score) where score 0–10.
    """
    text = f"{domain} {title} {description}".lower()
    best_topic = "Unknown"
    best_score = 0

    for topic, keywords in TOPIC_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in text)
        if score > best_score:
            best_score = score
            best_topic = topic

    return best_topic, min(best_score, 10)


def fetch_meta(domain):
    """Fetch homepage title and meta description for topic scoring."""
    try:
        r = requests.get(
            f"https://{domain}",
            headers=HEADERS,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True
        )
        if r.status_code == 200:
            soup = BeautifulSoup(r.text[:8000], "lxml")
            title = soup.title.string.strip() if soup.title else ""
            meta  = ""
            meta_tag = soup.find("meta", attrs={"name": re.compile("description", re.I)})
            if meta_tag:
                meta = meta_tag.get("content", "")
            return title, meta
    except Exception:
        pass
    return "", ""


# ══════════════════════════════════════════════════════════════════════════════
# EMAIL SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

def clean_email(e):
    e = e.strip().lower()
    domain = e.split("@")[-1]
    if domain in NOISE_EMAIL_DOMAINS or "." not in domain:
        return None
    if any(e.endswith(x) for x in [".png",".jpg",".gif",".svg",".css",".js",".woff",".ttf"]):
        return None
    if len(e) > 80:
        return None
    return e


def fetch_page(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code == 200 and "text/html" in r.headers.get("Content-Type", ""):
            return BeautifulSoup(r.text, "lxml"), r.url
    except Exception:
        pass
    return None, None


def extract_contacts(soup):
    text    = soup.get_text(" ")
    raw_em  = set(EMAIL_RE.findall(text))
    for tag in soup.find_all("a", href=re.compile(r"^mailto:", re.I)):
        raw_em.add(tag["href"].replace("mailto:", "").split("?")[0])
    emails  = {clean_email(e) for e in raw_em} - {None}
    phones  = {p.strip() for p in PHONE_RE.findall(text) if len(re.sub(r"\D", "", p)) >= 9}
    return emails, phones


def classify_opportunity(pages_hit):
    joined = " ".join(pages_hit)
    if any(k in joined for k in ["advertis", "sponsor"]):
        return "Paid Placement"
    elif any(k in joined for k in ["write", "contribut"]):
        return "Guest Post"
    elif "partner" in joined:
        return "Partnership"
    elif "contact" in joined or "about" in joined:
        return "Editorial / Contact"
    return "Unknown"


def scrape_domain(entry):
    domain = entry["domain"]
    base   = f"https://{domain}"

    # Score topic relevance via homepage meta
    title, meta = fetch_meta(domain)
    topic, relevance_score = score_topic(domain, title, meta)
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    all_emails, all_phones = set(), set()
    pages_hit = []

    # Scrape contact paths
    for path in CONTACT_PATHS:
        soup, _ = fetch_page(base + path)
        if soup:
            e, p = extract_contacts(soup)
            if e or p:
                pages_hit.append(path)
                all_emails |= e
                all_phones |= p
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    return {
        "domain":             domain,
        "domain_rating":      entry["domain_rating"],
        "monthly_traffic":    entry["traffic_domain"],
        "topic_category":     topic,
        "relevance_score":    relevance_score,
        "competitors_on":     ", ".join(entry.get("competitors_found_on", [])),
        "emails":             "; ".join(sorted(all_emails)),
        "best_email":         sorted(all_emails)[0] if all_emails else "",
        "phones":             "; ".join(sorted(all_phones)),
        "pages_hit":          ", ".join(pages_hit) or "none",
        "opportunity_type":   classify_opportunity(pages_hit),
        "has_email":          bool(all_emails),
        "outreach_status":    "Not Started",
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(results, filepath):
    wb = Workbook()

    # ── Sheet 1: Full list ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Outreach List"

    navy   = "1F3864"
    blue   = "2E75B6"
    lblue  = "EBF3FB"
    green  = "E2EFDA"
    amber  = "FFF2CC"
    red    = "FCE4D6"
    grey   = "F2F2F2"

    def hdr_cell(cell, val, bg=navy, fg="FFFFFF", bold=True, size=10, center=False):
        cell.value = val
        cell.font  = Font(name="Arial", bold=bold, color=fg, size=size)
        cell.fill  = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=True
        )
        thin = Side(style="thin", color="D9D9D9")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def body_cell(cell, val, bg="FFFFFF", bold=False, color="000000", center=False):
        cell.value = val
        cell.font  = Font(name="Arial", size=9, bold=bold, color=color)
        cell.fill  = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=True
        )
        thin = Side(style="thin", color="E0E0E0")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"32red.com — Backlink Outreach List  |  Generated {datetime.now().strftime('%d %b %Y')}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=navy)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Total domains: {len(results)}  |  With email: {sum(1 for r in results if r['has_email'])}  |  Topics: Gambling, Sports, eSports, Finance, Tech, Entertainment, General News, Travel, Crypto, Magazines"
    ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="808080")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15

    headers = [
        "Domain", "DR", "Monthly Traffic", "Topic", "Relevance",
        "Best Email", "All Emails", "Phone",
        "Opportunity Type", "Competitors Linking", "Pages Hit", "Status"
    ]
    for c, h in enumerate(headers, 1):
        hdr_cell(ws.cell(3, c), h, center=(c in [2,3,5,9,12]))
    ws.row_dimensions[3].height = 20

    opp_fill = {
        "Paid Placement":        amber,
        "Guest Post":            green,
        "Partnership":           green,
        "Editorial / Contact":   lblue,
        "Unknown":               grey,
    }

    for i, row in enumerate(results):
        r = i + 4
        alt = lblue if i % 2 == 0 else "FFFFFF"

        # Domain
        body_cell(ws.cell(r,1), row["domain"], bold=True, color=navy)
        # DR with colour scale
        dr = row["domain_rating"]
        dr_bg = "375623" if dr >= 75 else ("70AD47" if dr >= 50 else "A9D18E")
        dr_fg = "FFFFFF" if dr >= 50 else "375623"
        body_cell(ws.cell(r,2), dr, bg=dr_bg, bold=True, color=dr_fg, center=True)
        # Traffic
        body_cell(ws.cell(r,3), row["monthly_traffic"], bg=alt, center=True)
        # Topic
        body_cell(ws.cell(r,4), row["topic_category"], bg=alt)
        # Relevance score
        rel = row["relevance_score"]
        rel_bg = green if rel >= 3 else (amber if rel >= 1 else red)
        body_cell(ws.cell(r,5), f"{rel}/10", bg=rel_bg, center=True)
        # Best email
        body_cell(ws.cell(r,6), row["best_email"], bg=alt, color=navy)
        # All emails
        body_cell(ws.cell(r,7), row["emails"], bg=alt)
        # Phone
        body_cell(ws.cell(r,8), row["phones"], bg=alt)
        # Opportunity type
        opp = row["opportunity_type"]
        body_cell(ws.cell(r,9), opp, bg=opp_fill.get(opp, grey), center=True)
        # Competitors
        body_cell(ws.cell(r,10), row["competitors_on"], bg=alt)
        # Pages
        body_cell(ws.cell(r,11), row["pages_hit"], bg=alt)
        # Status
        body_cell(ws.cell(r,12), row["outreach_status"], bg=grey, center=True)

        ws.row_dimensions[r].height = 28

    # Column widths
    for col, w in zip(range(1,13), [26,6,14,18,9,32,52,18,20,36,28,14]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze pane under header
    ws.freeze_panes = "A4"

    # ── Sheet 2: Emails Only (clean list) ──────────────────────────────────
    ws2 = wb.create_sheet("Emails Only")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Clean Email List — Domains with confirmed contact email"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=12, color=navy)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 24

    for c, h in enumerate(["Domain", "Best Email", "Topic", "DR"], 1):
        hdr_cell(ws2.cell(2, c), h)
    ws2.row_dimensions[2].height = 18

    email_rows = [r for r in results if r["has_email"]]
    for i, row in enumerate(email_rows):
        r = i + 3
        alt = lblue if i % 2 == 0 else "FFFFFF"
        body_cell(ws2.cell(r,1), row["domain"], bold=True, color=navy)
        body_cell(ws2.cell(r,2), row["best_email"], bg=alt, color=navy)
        body_cell(ws2.cell(r,3), row["topic_category"], bg=alt)
        body_cell(ws2.cell(r,4), row["domain_rating"], bg=alt, center=True)
        ws2.row_dimensions[r].height = 20

    for col, w in zip(range(1,5), [28, 38, 20, 8]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Summary ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Pipeline Summary"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color=navy)
    ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 26
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 14

    summary = [
        ("Total domains analysed", len(results)),
        ("Domains with email found", sum(1 for r in results if r["has_email"])),
        ("Contact form / no email", sum(1 for r in results if not r["has_email"])),
        ("", ""),
        ("By Opportunity Type", ""),
        ("Paid Placement", sum(1 for r in results if r["opportunity_type"]=="Paid Placement")),
        ("Guest Post", sum(1 for r in results if r["opportunity_type"]=="Guest Post")),
        ("Partnership", sum(1 for r in results if r["opportunity_type"]=="Partnership")),
        ("Editorial / Contact", sum(1 for r in results if r["opportunity_type"]=="Editorial / Contact")),
        ("Unknown", sum(1 for r in results if r["opportunity_type"]=="Unknown")),
        ("", ""),
        ("By Topic Category", ""),
    ]
    from collections import Counter
    topic_counts = Counter(r["topic_category"] for r in results)
    for topic, count in topic_counts.most_common():
        summary.append((topic, count))

    for row_idx, (label, val) in enumerate(summary, 2):
        c1 = ws3.cell(row_idx, 1, label)
        c2 = ws3.cell(row_idx, 2, val)
        is_section = label in ["By Opportunity Type", "By Topic Category"]
        is_total   = label == "Total domains analysed"
        c1.font = Font(name="Arial", size=9, bold=(is_section or is_total))
        c2.font = Font(name="Arial", size=9, bold=is_total)
        if is_section:
            c1.fill = PatternFill("solid", start_color=navy)
            c1.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            ws3.merge_cells(f"A{row_idx}:B{row_idx}")
        elif is_total:
            c1.fill = PatternFill("solid", start_color=blue)
            c1.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c2.fill = PatternFill("solid", start_color=blue)
            c2.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        ws3.row_dimensions[row_idx].height = 16

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    wb.save(filepath)
    print(f"\n  ✓ Excel saved → {filepath}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    if not AHREFS_API_KEY:
        raise ValueError("AHREFS_API_KEY environment variable not set.")

    print("=" * 60)
    print("  32red.com Outreach Pipeline")
    print(f"  {datetime.now().strftime('%A %d %B %Y, %H:%M')}")
    print("=" * 60)

    # Step 1–2: Gap analysis via Ahrefs
    gap_domains = build_gap_list()

    # Step 4: Scrape emails concurrently
    print(f"\n[4/5] Scraping {len(gap_domains)} domains for contact data...")
    results = []
    with ThreadPoolExecutor(max_workers=SCRAPE_WORKERS) as pool:
        futures = {pool.submit(scrape_domain, d): d["domain"] for d in gap_domains}
        done = 0
        for future in as_completed(futures):
            done += 1
            domain = futures[future]
            try:
                result = futures[future] = future.result()
                results.append(result)
                em = "✓" if result["has_email"] else "–"
                print(f"  [{done:>4}/{len(gap_domains)}] {em} {domain:35s}  {result['topic_category']}")
            except Exception as ex:
                print(f"  [{done:>4}/{len(gap_domains)}] ✗ {domain}: {ex}")

    # Sort: has email first, then by DR
    results.sort(key=lambda x: (-int(x["has_email"]), -x["domain_rating"]))

    # Step 5: Write Excel
    print(f"\n[5/5] Writing Excel output...")
    date_str  = datetime.now().strftime("%Y-%m-%d")
    out_path  = os.path.join(OUTPUT_DIR, f"32red_outreach_{date_str}.xlsx")
    write_excel(results, out_path)

    # Summary
    with_email = sum(1 for r in results if r["has_email"])
    print(f"\n{'=' * 60}")
    print(f"  DONE  |  {len(results)} domains  |  {with_email} with email")
    print(f"  File  →  {out_path}")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
